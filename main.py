import streamlit as st
import PyPDF2
import docx
import csv
import openpyxl
import os
from groq import Groqclient
from dotenv import load_dotenv
 
# Load environment variables
load_dotenv()
groq_api_key = os.getenv("GROQ_API_KEY")
 
# Initialize Groq client
client = Groqclient(api_key=groq_api_key)
 
# Function to generate response using the Groq model
def generate_response(prompt):
    try:
        completion = client.chat.completions.create(
            model="mixtral-8x7b-32768",
            messages=[{"role": "user", "content": prompt}],
            temperature=1,
            max_tokens=1024,
            top_p=1,
            stream=True
        )
        response_text = ""
        for chunk in completion:
            response_text += chunk.choices[0].delta.content or ""
        return response_text
    except Exception as e:
        st.error(f"Error generating response: {e}")
        return ""
 
# Functions to extract text from files
def extract_text_from_pdf(file):
    pdf_reader = PyPDF2.PdfReader(file)
    text = ''
    for page in pdf_reader.pages:
        text += page.extract_text() or ''
    return text
 
def extract_text_from_docx(file):
    doc = docx.Document(file)
    text = ''
    for para in doc.paragraphs:
        text += para.text + '\n'
    return text
 
def extract_text_from_csv(file):
    text = ''
    reader = csv.reader(file)
    for row in reader:
        text += ', '.join(row) + '\n'
    return text
 
def extract_text_from_excel(file):
    wb = openpyxl.load_workbook(file)
    text = ''
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        for row in ws.iter_rows(values_only=True):
            text += ', '.join(str(cell) if cell is not None else '' for cell in row) + '\n'
    return text
 
def extract_text(file):
    if file.name.endswith(".pdf"):
        return extract_text_from_pdf(file)
    elif file.name.endswith(".docx"):
        return extract_text_from_docx(file)
    elif file.name.endswith(".csv"):
        return extract_text_from_csv(file)
    elif file.name.endswith(".xlsx") or file.name.endswith(".xls"):
        return extract_text_from_excel(file)
    else:
        return "Unsupported file format."
 
# Function to chunk large text
def chunk_large_text(text, chunk_size=3000):
    return [text[i:i+chunk_size] for i in range(0, len(text), chunk_size)]
 
# Streamlit UI
st.title("KITE.AI")
 
# File upload
uploaded_file = st.file_uploader("Upload a file", type=["pdf", "docx", "csv", "xlsx", "xls"])
 
if uploaded_file:
    with st.spinner(f"Processing {uploaded_file.name}..."):
        file_text = extract_text(uploaded_file)
   
    if file_text == "Unsupported file format.":
        st.error(file_text)
    else:
        st.success(f"{uploaded_file.name} uploaded successfully!")
       
        # Chatbot interface
        st.write("### Chat with KITE.AI about the uploaded document")
        if "chat_history" not in st.session_state:
            st.session_state.chat_history = []
       
        # User input
        user_input = st.text_input("Ask KITE.AI anything about the document")
       
        if user_input:
            st.session_state.chat_history.append(f"You: {user_input}")
           
            # Check if document exceeds token limit and chunk if needed
            chunks = chunk_large_text(file_text)
            all_responses = []
 
            for chunk in chunks:
                prompt = f"""
                You are a testing assistant. Based on the following document text, generate a detailed response according to the user's query.
 
                User Query: {user_input}
 
                Document Text:
                {chunk}
 
                Please format the response as follows:
                - For Test Cases: Module Name, Scenario, Flow (Positive/Negative), Business Priority (High/Low/Medium).
                - For End-to-End Scenarios: Module Name, Scenario, Flow (Positive/Negative), Business Priority (High/Low/Medium), Test Case Description, Test Steps, Expected Results.
                - For UI Specific Scenarios: Module Name, Scenario, Flow (Positive/Negative), Business Priority (High/Low/Medium), Test Case Description, Test Steps, Expected Results.
                - For Boundary Value Scenarios: Module Name, Scenario, Flow (Positive/Negative), Business Priority (High/Low/Medium), Test Case Description, Test Steps, Expected Results.
                - For Equivalence Partitioning Scenarios: Module Name, Scenario, Flow (Positive/Negative), Business Priority (High/Low/Medium), Test Case Description, Test Steps, Expected Results.
                - For Critical Scenarios: Module Name, Scenario, Flow (Positive/Negative), Business Priority (High/Low/Medium), Test Case Description, Test Steps, Expected Results.
                - For Test Scenarios: Module Name, Scenario, Flow (Positive/Negative), Business Priority (High/Low/Medium), Test Case Description, Test Steps, Expected Results.
                """
               
                try:
                    with st.spinner("KITE.AI is generating..."):
                        response = generate_response(prompt)
                        all_responses.append(response)
                except Exception as e:
                    st.error(f"Error generating response: {e}")
                    break
           
            # Combine all responses
            full_response = " ".join(all_responses)
            st.session_state.chat_history.append(f"KITE.AI: {full_response}")
       
        # Display chat history
        st.write("### Chat History")
        for message in st.session_state.chat_history:
            st.write(message)
else:
    st.info("Please upload a file to start the conversation.")
 